import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from weasyprint import HTML

# ---------------------------------------------------------
# 1. Excel Data Storage
# ---------------------------------------------------------
excel_filename = "customer_payments.xlsx"

# Header columns
headers = [
    "ID", "Status", "Date", "Corporation", "Applicant Name", "Father/Husband Name",
    "Address", "Aadhaar", "Gender", "Religion", "Caste", "Bank Name", "Branch",
    "Account No", "IFSC Code", "Loan Required", "Disabled", "Marital Status", "Mobile",
    "Project Name", "BHK/Unit", "Flat No", "Total Cost", "MHADA Form Fees", "Doc Fees",
    "Received Amount", "Balance Amount", "Payment Status", "Installment 1", "Installment 2"
]

row_data = [
    "MK-00001", "Old Active", "18-Jan-2023", "म्हाडा (MHADA)", "Pratiksha Gajanan Suryavanshi", "Gajanan Suryawanshi",
    "Flat No 402 Suyog Building Lane No 1, Mangal Nagar, Datta Mandir Road, Thergaon, Pin - 411057",
    "[Aadhaar Redacted]", "स्त्री (०२) / Female (02)", "Hindu (01)", "01 (General)", "HDFC BANK", "Thergaon",
    "50100323760260", "HDFC0002150", "No", "No", "01 (Married)", "9689929009",
    "VTP", "1BHK", "Flat", 1100000, 362800, 40000, 362800, 0, "Fully Paid", 62800, 300000
]

if not os.path.exists(excel_filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment Records"
    
    # Style Header
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
else:
    wb = openpyxl.load_workbook(excel_filename)
    ws = wb.active

ws.append(row_data)

# Adjust column widths
for col in ws.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = openpyxl.utils.get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

wb.save(excel_filename)

# ---------------------------------------------------------
# 2. HTML to PDF Receipt Generation
# ---------------------------------------------------------
html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>Payment Receipt - MK-00001</title>
    <style>
        @page {{
            size: A4;
            margin: 15mm 12mm;
            background-color: #faf8f5;
        }}
        * {{
            box-sizing: border-box;
        }}
        body {{
            font-family: Arial, sans-serif;
            margin: 0;
            padding: 0;
            color: #2b2b2b;
            font-size: 10pt;
        }}
        .header {{
            background-color: #1f4e79;
            color: #ffffff;
            padding: 20px;
            margin: -15mm -12mm 15mm -12mm;
            text-align: center;
        }}
        .header h1 {{
            margin: 0;
            font-size: 20pt;
            letter-spacing: 1px;
        }}
        .header p {{
            margin: 5px 0 0 0;
            font-size: 11pt;
            opacity: 0.9;
        }}
        .receipt-badge {{
            display: inline-block;
            background-color: #27ae60;
            color: white;
            padding: 5px 15px;
            border-radius: 12px;
            font-weight: bold;
            font-size: 11pt;
            margin-top: 10px;
        }}
        .details-box {{
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 20px;
            background-color: #ffffff;
            border: 1px solid #e0e0e0;
        }}
        .details-box th, .details-box td {{
            padding: 9px 12px;
            border: 1px solid #e0e0e0;
            text-align: left;
        }}
        .details-box th {{
            background-color: #f2f4f7;
            color: #1f4e79;
            width: 30%;
            font-weight: bold;
        }}
        .section-title {{
            font-size: 12pt;
            color: #1f4e79;
            border-left: 4px solid #1f4e79;
            padding-left: 8px;
            margin: 15px 0 10px 0;
            font-weight: bold;
        }}
        .summary-box {{
            background-color: #eaf2f8;
            border: 1px solid #b8d1e5;
            padding: 12px;
            margin-top: 15px;
            text-align: center;
        }}
        .summary-box h2 {{
            margin: 0;
            color: #27ae60;
            font-size: 16pt;
        }}
        .footer {{
            margin-top: 30px;
            text-align: right;
            padding-right: 10px;
        }}
        .footer-text {{
            font-size: 9pt;
            color: #7f8c8d;
            text-align: center;
            margin-top: 40px;
            border-top: 1px solid #e0e0e0;
            padding-top: 10px;
        }}
    </style>
</head>
<body>

    <div class="header">
        <h1>M K ENTERPRISES</h1>
        <p>Thergaon, Wakad Road, Pune - 411057</p>
        <div class="receipt-badge">पेमेंट पावती / PAYMENT RECEIPT</div>
    </div>

    <table class="details-box">
        <tr>
            <th>पावती क्र. / Receipt No:</th>
            <td>MK-00001</td>
            <th>दिनांक / Date:</th>
            <td>18-Jan-2023</td>
        </tr>
        <tr>
            <th>ग्राहक नाव / Name:</th>
            <td colspan="3"><strong>Pratiksha Gajanan Suryavanshi</strong></td>
        </tr>
        <tr>
            <th>पती/वडिलांचे नाव / Husband's Name:</th>
            <td colspan="3">Gajanan Suryawanshi</td>
        </tr>
        <tr>
            <th>पत्ता / Address:</th>
            <td colspan="3">Flat No 402 Suyog Building Lane No 1, Mangal Nagar, Datta Mandir Road, Thergaon, Pin - 411057</td>
        </tr>
        <tr>
            <th>मोबाईल क्र. / Mobile:</th>
            <td>9689929009</td>
            <th>योजना / Scheme:</th>
            <td>म्हाडा (MHADA) 2023</td>
        </tr>
    </table>

    <div class="section-title">फ्लॅट व फॉर्म तपशील / Flat & Application Details</div>
    <table class="details-box">
        <tr>
            <th>प्रकल्पाचे नाव / Project:</th>
            <td>VTP</td>
            <th>प्रकार / Unit:</th>
            <td>1BHK (Flat)</td>
        </tr>
        <tr>
            <th>फ्लॅटची एकूण किंमत / Total Cost:</th>
            <td>₹ 11,000,000</td>
            <th>फॉर्म फी / Form Fees:</th>
            <td>₹ 362,800</td>
        </tr>
    </table>

    <div class="section-title">पेमेंट तपशील / Payment Breakdown</div>
    <table class="details-box">
        <thead>
            <tr style="background-color: #f2f4f7;">
                <th style="width: 50%;">तपशील / Description</th>
                <th style="width: 50%;">रक्कम / Amount (₹)</th>
            </tr>
        </thead>
        <tbody>
            <tr>
                <td>हप्ता १ / Installment 1</td>
                <td>₹ 62,800</td>
            </tr>
            <tr>
                <td>हप्ता २ / Installment 2</td>
                <td>₹ 300,000</td>
            </tr>
            <tr style="font-weight: bold; background-color: #f9f9f9;">
                <td>एकूण प्राप्त रक्कम / Total Received:</td>
                <td>₹ 362,800</td>
            </tr>
            <tr style="font-weight: bold; background-color: #f9f9f9;">
                <td>उर्वरित रक्कम / Balance Amount:</td>
                <td>₹ 0 (Nil)</td>
            </tr>
        </tbody>
    </table>

    <div class="summary-box">
        <span>पेमेंट स्थिती / Payment Status:</span>
        <h2>✅ Fully Paid (संपूर्ण जमा)</h2>
    </div>

    <div class="footer">
        <p><strong>M K ENTERPRISES साठी</strong></p>
        <br><br>
        <p>(अधिकृत स्वाक्षरी / Authorized Signatory)</p>
    </div>

    <div class="footer-text">
        आपल्या पेमेंटची नोंद यशस्वीरीत्या पूर्ण झाली आहे. सहकार्याबद्दल धन्यवाद! 🙏
    </div>

</body>
</html>
"""

with open("receipt_MK-00001.html", "w", encoding="utf-8") as f:
    f.write(html_content)

HTML("receipt_MK-00001.html").write_pdf("Payment_Receipt_Pratiksha_Suryavanshi.pdf")